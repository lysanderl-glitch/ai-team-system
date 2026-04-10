"""
WBS Excel -> JSON export for Notion import.
Reads Janusd WBS交付 Excel and outputs structured JSON.
"""

import json
import re
import sys
from pathlib import Path

import openpyxl

# --- File resolution ---
CANDIDATES = [
    Path(r"C:\Users\lysanderl_janusd\PMO-AI Auto\PMO-AI Auto\Janusd_WBS_交付_V3.0_Final.xlsx"),
    Path(r"C:\Users\lysanderl_janusd\Downloads\Janusd_WBS_交付_审查版.xlsx"),
]

excel_path = None
for p in CANDIDATES:
    if p.exists():
        excel_path = p
        break

if excel_path is None:
    # Fallback: glob for *WBS*交付*.xlsx
    fallback_dir = Path(r"C:\Users\lysanderl_janusd\PMO-AI Auto\PMO-AI Auto")
    matches = sorted(fallback_dir.glob("*WBS*交付*.xlsx"), key=lambda f: f.stat().st_mtime, reverse=True)
    if matches:
        excel_path = matches[0]

if excel_path is None:
    print("ERROR: No WBS Excel file found.")
    sys.exit(1)

print(f"Reading: {excel_path}")

# --- Mappings ---

STAGE_PREFIX_MAP = {
    "S": "S-售前",
    "DA": "DA-启动筹备",
    "DP": "DP-项目策划",
    "DO": "DO-开工会",
    "DC": "DC-沟通管理",
    "DD": "DD-深化设计",
    "DS": "DS-系统部署",
    "DY": "DY-系统调试",
    "DB": "DB-BIM建模",
    "DR": "DR-BIM审核",
    "DI": "DI-BIM集成",
    "DT": "DT-测试验收",
    "DU": "DU-培训移交",
    "DV": "DV-质保运维",
}

# 执行流 mapping
FLOW_MAP = {
    "S": "P-EX1", "DA": "P-EX1", "DP": "P-EX1", "DO": "P-EX1",
    "DC": "P-EX1", "DT": "P-EX1", "DU": "P-EX1", "DV": "P-EX1",
    "DD": "P-EX2",
    "DS": "P-EX3",  # except DS007/DS008
    "DY": "P-EX4",
    "DB": "P-EX5", "DR": "P-EX5", "DI": "P-EX5",
}

STAGE_ROLE_MAP = {
    "S": ["Sales", "SA"],
    "DA": ["PM", "DE"],
    "DP": ["PM", "SA", "CDE"],
    "DO": ["PM"],
    "DC": ["PM"],
    "DD": ["SA", "PM"],
    "DS": ["DE"],
    "DY": ["DE"],
    "DB": ["CDE"],
    "DR": ["CDE"],
    "DI": ["CDE", "DE"],
    "DT": ["QA", "PM", "CDE"],
    "DU": ["PM", "CDE"],
    "DV": ["PM", "Sales"],
}

GATE_MAP = {
    "S": "G0",
    "DA": "G1", "DP": "G1",
    "DS": "G2", "DY": "G2",
    "DB": "G3", "DR": "G3", "DI": "G3", "DD": "G3",
    "DT": "G4", "DU": "G4", "DV": "G4",
    "DO": "G1", "DC": "G1",  # reasonable defaults
}


def extract_prefix(wbs_code: str) -> str:
    """Extract stage prefix from WBS code like S001, DA001, DP001001."""
    m = re.match(r"^([A-Z]+)", str(wbs_code).strip())
    return m.group(1) if m else ""


def get_flow(wbs_code: str, prefix: str) -> str:
    code = str(wbs_code).strip()
    # DS007, DS008 -> P-EX6
    if prefix == "DS" and re.match(r"DS0*0?[78]", code):
        return "P-EX6"
    return FLOW_MAP.get(prefix, "")


# --- Read Excel ---
wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)

# Try to find the sheet
target_sheet = None
for name in wb.sheetnames:
    if "WBS" in name and "主表" in name:
        target_sheet = name
        break
if target_sheet is None:
    target_sheet = wb.sheetnames[0]

print(f"Sheet: {target_sheet}")
ws = wb[target_sheet]

tasks = []
errors = []
level_counts = {1: 0, 2: 0, 3: 0, 4: 0}

for row_idx, row in enumerate(ws.iter_rows(min_row=4, max_col=6, values_only=True), start=4):
    wbs_code = row[0]
    if wbs_code is None or str(wbs_code).strip() == "":
        continue

    wbs_code = str(wbs_code).strip()
    level_raw = row[1]
    task_name = row[2]
    duration = row[3]
    parallel_group = row[4]
    dependency = row[5]

    # Parse level
    try:
        level = int(level_raw) if level_raw is not None else 0
    except (ValueError, TypeError):
        level = 0
        errors.append(f"Row {row_idx}: invalid level '{level_raw}' for {wbs_code}")

    if level in level_counts:
        level_counts[level] += 1

    prefix = extract_prefix(wbs_code)
    stage = STAGE_PREFIX_MAP.get(prefix, prefix)
    flow = get_flow(wbs_code, prefix)
    roles = STAGE_ROLE_MAP.get(prefix, [])
    gate = GATE_MAP.get(prefix, "")

    if not prefix:
        errors.append(f"Row {row_idx}: cannot parse prefix from '{wbs_code}'")

    task = {
        "序号": row_idx - 3,
        "WBS编码": wbs_code,
        "层级": level,
        "任务名称": str(task_name).strip() if task_name else "",
        "工期(天)": float(duration) if duration is not None else None,
        "并行组": str(parallel_group).strip() if parallel_group else "",
        "前置依赖": str(dependency).strip() if dependency else "",
        "负责角色": roles,
        "所属阶段": stage,
        "执行流": flow,
        "阶段门": gate,
    }
    tasks.append(task)

wb.close()

# --- Output ---
output_path = Path(r"C:\Users\lysanderl_janusd\Claude Code\ai-team-system\scripts\_wbs_export.json")
with open(output_path, "w", encoding="utf-8") as f:
    json.dump(tasks, f, ensure_ascii=False, indent=2)

# --- Report ---
print(f"\nTotal tasks: {len(tasks)}")
print(f"Breakdown by level:")
for lvl in sorted(level_counts.keys()):
    print(f"  L{lvl}: {level_counts[lvl]}")
if errors:
    print(f"\nErrors ({len(errors)}):")
    for e in errors:
        print(f"  {e}")
else:
    print("\nNo errors.")
print(f"\nJSON written to: {output_path}")
