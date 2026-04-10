"""
WBS 跨流依赖完整性校验脚本
检查前置依赖链是否存在断链（依赖的WBS编码不存在）
由 janus_pmo_auto 负责运维
"""
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    import openpyxl
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)


def check_dependencies(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb["① WBS主表"]

    # 第一遍：收集所有WBS编码
    all_codes = set()
    tasks = []

    for row in ws.iter_rows(min_row=4, values_only=False):
        wbs_code = row[0].value  # A列
        name = row[2].value      # C列
        deps = row[5].value      # F列: 前置依赖

        if wbs_code:
            code = str(wbs_code).strip()
            all_codes.add(code)
            if deps:
                tasks.append({
                    "wbs": code,
                    "name": name or "",
                    "deps_raw": str(deps).strip()
                })

    # 第二遍：校验依赖是否存在
    broken = []
    for task in tasks:
        dep_list = [d.strip() for d in task["deps_raw"].replace("，", ",").split(",")]
        for dep in dep_list:
            if dep and dep not in all_codes:
                broken.append({
                    "wbs": task["wbs"],
                    "name": task["name"],
                    "missing_dep": dep
                })

    # 第三遍：检查跨流依赖标记（Sheet⑤）
    cross_flow = []
    try:
        ws5 = wb["⑤ 参考手册"]
        for row in ws5.iter_rows(min_row=33, values_only=False):
            code = row[0].value
            desc = row[1].value
            detail = row[2].value
            if code and detail and "依赖" in str(detail):
                code_str = str(code).strip()
                if code_str in all_codes:
                    cross_flow.append({
                        "wbs": code_str,
                        "description": str(desc or ""),
                        "dependency_detail": str(detail or "")
                    })
    except (KeyError, Exception):
        pass

    return broken, cross_flow, len(all_codes), len(tasks)


def main():
    import os
    default_path = os.path.join(
        os.path.expanduser("~"),
        "Downloads",
        "Janusd_WBS_交付_审查版 (2).xlsx"
    )
    filepath = sys.argv[1] if len(sys.argv) > 1 else default_path

    if not os.path.exists(filepath):
        print(f"文件不存在: {filepath}")
        sys.exit(1)

    print(f"🔗 WBS 跨流依赖完整性校验")
    print(f"   文件: {os.path.basename(filepath)}")
    print("=" * 60)

    broken, cross_flow, total_codes, total_with_deps = check_dependencies(filepath)

    print(f"📊 统计: {total_codes} 个WBS编码, {total_with_deps} 个有前置依赖\n")

    if not broken:
        print("✅ 所有前置依赖引用均有效，无断链。")
    else:
        print(f"❌ 发现 {len(broken)} 个断链依赖：\n")
        for i, item in enumerate(broken, 1):
            print(f"  {i}. {item['wbs']} ({item['name']})")
            print(f"     → 依赖 [{item['missing_dep']}] 不存在于WBS编码表")
            print()

    if cross_flow:
        print(f"\n📌 跨流依赖标记 ({len(cross_flow)} 个)：\n")
        for item in cross_flow:
            print(f"  {item['wbs']} {item['description']}")
            print(f"     {item['dependency_detail']}")
            print()

    print("=" * 60)
    print(f"校验完成。断链: {len(broken)}, 跨流依赖: {len(cross_flow)}")
    return len(broken)


if __name__ == "__main__":
    sys.exit(0 if main() == 0 else 1)
