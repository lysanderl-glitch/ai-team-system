"""
WBS数据源抽象层
统一Excel和Notion两种后端的数据访问接口。
所有PMO脚本通过此模块读取WBS数据，无需关心底层存储。

配置方式（环境变量）：
  WBS_SOURCE=excel     — 使用Excel后端（默认）
  WBS_SOURCE=notion    — 使用Notion后端
  WBS_EXCEL_PATH=...   — Excel文件路径
  NOTION_TOKEN=...     — Notion API Token
  WBS_NOTION_DB_ID=... — Notion WBS数据库ID

由 janus_pmo_auto 负责运维
"""
import os
import sys
import io
from abc import ABC, abstractmethod

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Notion WBS数据库 data_source_id（新workspace lysanderl@janusd.io）
DEFAULT_NOTION_DB_ID = "8cf298a9-fee0-4194-b371-b6616e2bb6aa"

# WBS阶段到角色的默认映射（共享，避免各脚本重复定义）
STAGE_ROLE_MAP = {
    "S":  ["Sales", "SA"],
    "DA": ["PM", "DE"],
    "DP": ["PM", "SA", "CDE"],
    "DO": ["PM"],
    "DC": ["PM"],
    "DD": ["SA", "PM"],
    "DS": ["DE"],
    "DY": ["DE"],
    "DB": ["CDE"],
    "DR": ["CDE"],
    "DI": ["CDE", "DE"],
    "DT": ["QA", "PM", "CDE"],
    "DU": ["PM", "CDE"],
    "DV": ["PM", "Sales"],
}

# 阶段到执行流映射
STAGE_FLOW_MAP = {
    "S": "P-EX1", "DA": "P-EX1", "DP": "P-EX1", "DO": "P-EX1",
    "DC": "P-EX1", "DT": "P-EX1", "DU": "P-EX1", "DV": "P-EX1",
    "DD": "P-EX2",
    "DS": "P-EX3",
    "DY": "P-EX4",
    "DB": "P-EX5", "DR": "P-EX5", "DI": "P-EX5",
}

# 阶段到阶段门映射
STAGE_GATE_MAP = {
    "S": "G0", "DA": "G1", "DP": "G1",
    "DO": "G1", "DC": "G1",
    "DS": "G2", "DY": "G2",
    "DD": "G3", "DB": "G3", "DR": "G3", "DI": "G3",
    "DT": "G4", "DU": "G4", "DV": "G4",
}


def get_stage_prefix(wbs_code):
    """从WBS编码提取阶段前缀（如 DS001 → DS, S001 → S）"""
    prefix = ""
    for ch in wbs_code:
        if ch.isdigit():
            break
        prefix += ch
    return prefix


def get_roles_for_task(wbs_code):
    """根据WBS编码确定负责角色"""
    for prefix_len in range(len(wbs_code), 0, -1):
        prefix = wbs_code[:prefix_len]
        if prefix in STAGE_ROLE_MAP:
            return STAGE_ROLE_MAP[prefix]
    prefix = get_stage_prefix(wbs_code)
    return STAGE_ROLE_MAP.get(prefix, ["PM"])


# ---------------------------------------------------------------------------
# 抽象基类
# ---------------------------------------------------------------------------

class WBSDataSource(ABC):
    """WBS数据源抽象接口"""

    @abstractmethod
    def load_tasks(self):
        """
        加载WBS主表所有任务。
        返回 dict[str, dict]，key=WBS编码，value=任务字典：
        {
            "wbs": str,
            "level": float,
            "name": str,
            "duration": float,
            "parallel_group": str | None,
            "deps": list[str],
        }
        """
        pass

    @abstractmethod
    def load_team_config(self):
        """
        加载项目团队配置（Sheet③）。
        返回 list[dict]，每项 {"role": str, "person": str}
        """
        pass

    @abstractmethod
    def load_cross_flow_deps(self):
        """
        加载跨流依赖标记（Sheet⑤）。
        返回 list[dict]，每项 {"wbs": str, "description": str, "detail": str}
        """
        pass

    def get_source_name(self):
        """返回数据源标识"""
        return self.__class__.__name__


# ---------------------------------------------------------------------------
# Excel 后端
# ---------------------------------------------------------------------------

class ExcelWBSSource(WBSDataSource):
    """从Excel文件读取WBS数据"""

    def __init__(self, filepath):
        self.filepath = filepath
        try:
            import openpyxl
            self._openpyxl = openpyxl
        except ImportError:
            print("Excel后端需要安装 openpyxl: pip install openpyxl")
            sys.exit(1)

    def load_tasks(self):
        wb = self._openpyxl.load_workbook(self.filepath)
        ws = wb["① WBS主表"]

        tasks = {}
        for row in ws.iter_rows(min_row=4, values_only=False):
            wbs_code = row[0].value
            level = row[1].value
            name = row[2].value
            duration = row[3].value
            parallel_group = row[4].value
            dependency = row[5].value

            if wbs_code is None or level is None:
                continue

            try:
                level_num = float(level)
                dur = float(duration) if duration else 0
            except (ValueError, TypeError):
                continue

            deps = []
            if dependency:
                deps = [d.strip() for d in str(dependency).replace("，", ",").split(",")]
                deps = [d for d in deps if d]

            pg = str(parallel_group).strip() if parallel_group else None

            tasks[str(wbs_code).strip()] = {
                "wbs": str(wbs_code).strip(),
                "level": level_num,
                "name": str(name).strip() if name else "",
                "duration": dur,
                "parallel_group": pg,
                "deps": deps,
            }

        return tasks

    def load_team_config(self):
        wb = self._openpyxl.load_workbook(self.filepath)
        try:
            ws = wb["③ 项目团队配置"]
        except KeyError:
            return []

        config = []
        for row in ws.iter_rows(min_row=2, values_only=False):
            role = row[0].value
            person = row[1].value
            if role:
                config.append({
                    "role": str(role).strip(),
                    "person": str(person).strip() if person else "",
                })
        return config

    def load_cross_flow_deps(self):
        wb = self._openpyxl.load_workbook(self.filepath)
        try:
            ws = wb["⑤ 参考手册"]
        except KeyError:
            return []

        cross_flow = []
        for row in ws.iter_rows(min_row=33, values_only=False):
            code = row[0].value
            desc = row[1].value
            detail = row[2].value
            if code and detail and "依赖" in str(detail):
                cross_flow.append({
                    "wbs": str(code).strip(),
                    "description": str(desc).strip() if desc else "",
                    "detail": str(detail).strip(),
                })
        return cross_flow

    def get_source_name(self):
        return f"Excel({os.path.basename(self.filepath)})"


# ---------------------------------------------------------------------------
# Notion 后端
# ---------------------------------------------------------------------------

class NotionWBSSource(WBSDataSource):
    """从Notion数据库读取WBS数据"""

    NOTION_BASE = "https://api.notion.com/v1"
    NOTION_VERSION = "2022-06-28"

    # Notion select值到内部值的映射
    _PARALLEL_GROUP_MAP = {
        "P-TR": "P-TR", "P-FA": "P-FA", "P-IOT": "P-IOT",
        "P-BIZ": "P-BIZ", "P-MEP": "P-MEP", "无": None,
    }

    def __init__(self, data_source_id=None, token=None):
        self.data_source_id = data_source_id or DEFAULT_NOTION_DB_ID
        self.token = token or os.environ.get("NOTION_TOKEN", "")
        if not self.token:
            print("Notion后端需要设置 NOTION_TOKEN 环境变量")
            sys.exit(1)
        try:
            import requests
            self._requests = requests
        except ImportError:
            print("Notion后端需要安装 requests: pip install requests")
            sys.exit(1)

    def _headers(self):
        return {
            "Authorization": f"Bearer {self.token}",
            "Content-Type": "application/json",
            "Notion-Version": self.NOTION_VERSION,
        }

    def _query_all_pages(self, database_id=None):
        """查询Notion数据库所有页面（自动分页）"""
        db_id = database_id or self.data_source_id
        url = f"{self.NOTION_BASE}/databases/{db_id}/query"
        pages = []
        payload = {"page_size": 100}

        while True:
            resp = self._requests.post(url, headers=self._headers(), json=payload)
            resp.raise_for_status()
            data = resp.json()
            pages.extend(data.get("results", []))
            if not data.get("has_more"):
                break
            payload["start_cursor"] = data["next_cursor"]

        return pages

    def _extract_title(self, props, key="WBS编码"):
        """提取Title属性值"""
        title_prop = props.get(key, {})
        title_list = title_prop.get("title", [])
        if title_list:
            return title_list[0].get("plain_text", "").strip()
        return ""

    def _extract_text(self, props, key):
        """提取Rich Text属性值"""
        text_prop = props.get(key, {})
        text_list = text_prop.get("rich_text", [])
        if text_list:
            return text_list[0].get("plain_text", "").strip()
        return ""

    def _extract_number(self, props, key):
        """提取Number属性值"""
        num_prop = props.get(key, {})
        return num_prop.get("number")

    def _extract_select(self, props, key):
        """提取Select属性值"""
        sel_prop = props.get(key, {})
        sel = sel_prop.get("select")
        if sel:
            return sel.get("name", "")
        return ""

    def _extract_multi_select(self, props, key):
        """提取Multi-Select属性值"""
        ms_prop = props.get(key, {})
        ms_list = ms_prop.get("multi_select", [])
        return [item.get("name", "") for item in ms_list]

    def load_tasks(self):
        pages = self._query_all_pages()

        tasks = {}
        for page in pages:
            props = page.get("properties", {})

            wbs_code = self._extract_title(props, "WBS编码")
            if not wbs_code:
                continue

            level = self._extract_number(props, "层级")
            name = self._extract_text(props, "任务名称")
            duration = self._extract_number(props, "工期(天)")
            pg_raw = self._extract_select(props, "并行组")
            deps_raw = self._extract_text(props, "前置依赖")

            if level is None:
                continue

            # 解析并行组
            pg = self._PARALLEL_GROUP_MAP.get(pg_raw, pg_raw if pg_raw else None)

            # 解析依赖
            deps = []
            if deps_raw:
                deps = [d.strip() for d in deps_raw.replace("，", ",").split(",")]
                deps = [d for d in deps if d]

            tasks[wbs_code] = {
                "wbs": wbs_code,
                "level": float(level),
                "name": name,
                "duration": float(duration) if duration else 0,
                "parallel_group": pg,
                "deps": deps,
            }

        return tasks

    def load_team_config(self):
        # Phase C将创建独立团队配置DB并用Relation关联
        # 当前返回空列表，由调用方回退到默认映射
        return []

    def load_cross_flow_deps(self):
        # 跨流依赖信息已编码在前置依赖字段中
        # Notion后端不需要单独的Sheet⑤
        tasks = self.load_tasks()
        cross_flow = []
        for code, task in tasks.items():
            src_prefix = get_stage_prefix(code)
            for dep in task["deps"]:
                dep_prefix = get_stage_prefix(dep)
                if src_prefix != dep_prefix:
                    cross_flow.append({
                        "wbs": code,
                        "description": task["name"],
                        "detail": f"跨流依赖: {dep}({dep_prefix}) → {code}({src_prefix})",
                    })
        return cross_flow

    def get_source_name(self):
        return f"Notion(WBS-DB:{self.data_source_id[:8]}...)"


# ---------------------------------------------------------------------------
# 工厂函数
# ---------------------------------------------------------------------------

def get_wbs_source(filepath=None):
    """
    获取WBS数据源实例。

    优先级：
    1. 如果传入filepath且不为空 → Excel后端
    2. 如果环境变量 WBS_SOURCE=notion → Notion后端
    3. 如果环境变量 WBS_EXCEL_PATH 有值 → Excel后端
    4. 默认 → Notion后端（迁移完成后的默认行为）

    用法：
        source = get_wbs_source()          # 自动选择
        source = get_wbs_source("xxx.xlsx") # 强制Excel
        tasks = source.load_tasks()
    """
    # 显式传入filepath → Excel
    if filepath:
        return ExcelWBSSource(filepath)

    # 环境变量控制
    wbs_source = os.environ.get("WBS_SOURCE", "").lower()

    if wbs_source == "notion":
        db_id = os.environ.get("WBS_NOTION_DB_ID", DEFAULT_NOTION_DB_ID)
        return NotionWBSSource(data_source_id=db_id)

    if wbs_source == "excel":
        excel_path = os.environ.get("WBS_EXCEL_PATH", "")
        if not excel_path:
            print("WBS_SOURCE=excel 但未设置 WBS_EXCEL_PATH")
            sys.exit(1)
        return ExcelWBSSource(excel_path)

    # 检查WBS_EXCEL_PATH
    excel_path = os.environ.get("WBS_EXCEL_PATH", "")
    if excel_path and os.path.exists(excel_path):
        return ExcelWBSSource(excel_path)

    # 默认：Notion（Phase B验证通过后的最终状态）
    db_id = os.environ.get("WBS_NOTION_DB_ID", DEFAULT_NOTION_DB_ID)
    return NotionWBSSource(data_source_id=db_id)


# ---------------------------------------------------------------------------
# CLI — 用于验证数据源一致性
# ---------------------------------------------------------------------------

def main():
    """验证数据源并打印摘要"""
    import argparse
    parser = argparse.ArgumentParser(description="WBS数据源抽象层 — 验证与诊断")
    parser.add_argument("--excel", help="Excel文件路径")
    parser.add_argument("--notion", action="store_true", help="使用Notion后端")
    parser.add_argument("--compare", nargs=2, metavar=("EXCEL_PATH", "NOTION_DB_ID"),
                        help="对比Excel与Notion两个后端的数据一致性")
    args = parser.parse_args()

    if args.compare:
        excel_src = ExcelWBSSource(args.compare[0])
        notion_src = NotionWBSSource(data_source_id=args.compare[1])

        excel_tasks = excel_src.load_tasks()
        notion_tasks = notion_src.load_tasks()

        print(f"{'=' * 60}")
        print("📊 WBS数据源一致性对比")
        print(f"{'=' * 60}")
        print(f"  Excel任务数: {len(excel_tasks)}")
        print(f"  Notion任务数: {len(notion_tasks)}")

        # 编码对比
        excel_codes = set(excel_tasks.keys())
        notion_codes = set(notion_tasks.keys())

        only_excel = excel_codes - notion_codes
        only_notion = notion_codes - excel_codes
        common = excel_codes & notion_codes

        print(f"  共有编码: {len(common)}")
        if only_excel:
            print(f"  ⚠️ 仅Excel: {sorted(only_excel)[:10]}...")
        if only_notion:
            print(f"  ⚠️ 仅Notion: {sorted(only_notion)[:10]}...")

        # 字段对比
        mismatches = 0
        for code in sorted(common):
            et = excel_tasks[code]
            nt = notion_tasks[code]
            for field in ["level", "duration", "name", "parallel_group"]:
                ev = et.get(field)
                nv = nt.get(field)
                if field in ("level", "duration"):
                    if ev is not None and nv is not None and abs(float(ev) - float(nv)) > 0.01:
                        print(f"  ⚠️ {code}.{field}: Excel={ev} vs Notion={nv}")
                        mismatches += 1
                else:
                    if str(ev or "") != str(nv or ""):
                        mismatches += 1

        if mismatches == 0:
            print(f"\n  ✅ 数据完全一致")
        else:
            print(f"\n  ⚠️ 发现 {mismatches} 处差异")

        return 0

    # 单数据源验证
    if args.excel:
        source = ExcelWBSSource(args.excel)
    elif args.notion:
        source = NotionWBSSource()
    else:
        source = get_wbs_source()

    tasks = source.load_tasks()

    print(f"{'=' * 60}")
    print(f"📋 WBS数据源: {source.get_source_name()}")
    print(f"{'=' * 60}")
    print(f"  总任务数: {len(tasks)}")

    # 按层级统计
    by_level = {}
    for t in tasks.values():
        lv = int(t["level"])
        by_level[lv] = by_level.get(lv, 0) + 1
    for lv in sorted(by_level):
        print(f"  L{lv}: {by_level[lv]} 任务")

    # 按阶段统计
    by_stage = {}
    for code in tasks:
        prefix = get_stage_prefix(code)
        by_stage[prefix] = by_stage.get(prefix, 0) + 1
    print(f"\n  阶段分布:")
    for stage in sorted(by_stage):
        print(f"    {stage}: {by_stage[stage]}")

    # 依赖统计
    dep_count = sum(1 for t in tasks.values() if t["deps"])
    print(f"\n  有前置依赖的任务: {dep_count}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
