"""
WBS角色负载分析工具
基于WBS主表分析各角色(PM/DE/SA/CDE/Sales)的任务分布、工期负载、
关键路径暴露度，识别瓶颈角色和资源冲突风险。
由 janus_pmo_auto 负责运维
"""
import sys
import io
import os
from collections import defaultdict

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

try:
    import openpyxl
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)


# WBS阶段到角色的默认映射（基于janus_experts.yaml定义）
STAGE_ROLE_MAP = {
    "S":  ["Sales", "SA"],
    "DA": ["PM", "DE"],
    "DP": ["PM", "SA", "CDE"],
    "DO": ["PM"],
    "DC": ["PM"],
    "DD": ["SA", "PM"],
    "DS": ["DE"],
    "DY": ["DE"],
    "DB": ["CDE"],
    "DR": ["CDE"],
    "DI": ["CDE", "DE"],
    "DT": ["QA", "PM", "CDE"],
    "DU": ["PM", "CDE"],
    "DV": ["PM", "Sales"],
}


def get_roles_for_task(wbs_code):
    """根据WBS编码确定负责角色"""
    # 匹配最长前缀
    for prefix_len in range(len(wbs_code), 0, -1):
        prefix = wbs_code[:prefix_len]
        if prefix in STAGE_ROLE_MAP:
            return STAGE_ROLE_MAP[prefix]
    # 默认取L2前缀
    for i in range(len(wbs_code)):
        if wbs_code[i].isdigit():
            prefix = wbs_code[:i]
            if prefix in STAGE_ROLE_MAP:
                return STAGE_ROLE_MAP[prefix]
            break
    return ["PM"]  # 默认PM


def load_tasks(filepath):
    """加载WBS任务"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb["① WBS主表"]

    tasks = []
    for row in ws.iter_rows(min_row=4, values_only=False):
        wbs_code = row[0].value
        level = row[1].value
        name = row[2].value
        duration = row[3].value

        if wbs_code is None or level is None:
            continue

        try:
            level_num = float(level)
            dur = float(duration) if duration else 0
        except (ValueError, TypeError):
            continue

        # 只统计L3/L4实际执行任务
        if level_num >= 3:
            roles = get_roles_for_task(wbs_code)
            tasks.append({
                "wbs": wbs_code,
                "level": level_num,
                "name": str(name).strip() if name else "",
                "duration": dur,
                "roles": roles,
            })

    return tasks


def load_team_config(filepath):
    """加载项目团队配置（Sheet③）"""
    wb = openpyxl.load_workbook(filepath)
    try:
        ws = wb["③ 项目团队配置"]
    except KeyError:
        return {}

    team = {}
    for row in ws.iter_rows(min_row=2, values_only=False):
        role = row[0].value
        person = row[1].value
        if role and person:
            team[str(role).strip()] = str(person).strip()
    return team


def analyze_workload(tasks):
    """分析各角色工作负载"""
    role_stats = defaultdict(lambda: {
        "task_count": 0,
        "total_duration": 0,
        "l3_tasks": [],
        "l4_tasks": [],
        "stages": set(),
        "peak_concurrent": 0,
    })

    for task in tasks:
        for role in task["roles"]:
            stats = role_stats[role]
            stats["task_count"] += 1
            stats["total_duration"] += task["duration"]
            if task["level"] == 3:
                stats["l3_tasks"].append(task)
            else:
                stats["l4_tasks"].append(task)

            # 提取阶段
            stage = ""
            for i, ch in enumerate(task["wbs"]):
                if ch.isdigit():
                    stage = task["wbs"][:i]
                    break
            if stage:
                stats["stages"].add(stage)

    return dict(role_stats)


def main():
    default_path = os.path.join(
        os.path.expanduser("~"),
        "Downloads",
        "Janusd_WBS_交付_审查版 (2).xlsx"
    )
    filepath = sys.argv[1] if len(sys.argv) > 1 else default_path

    if not os.path.exists(filepath):
        print(f"文件不存在: {filepath}")
        sys.exit(1)

    print(f"👥 WBS角色负载分析")
    print(f"   文件: {os.path.basename(filepath)}")
    print("=" * 70)

    tasks = load_tasks(filepath)
    team = load_team_config(filepath)
    workload = analyze_workload(tasks)

    print(f"\n📊 总计 {len(tasks)} 个执行任务（L3+L4）")

    if team:
        print(f"\n👤 项目团队配置:")
        for role, person in team.items():
            print(f"    {role}: {person}")

    # 角色负载总览
    print(f"\n{'=' * 70}")
    print("📊 角色负载总览:")
    print(f"{'─' * 70}")
    print(f"{'角色':<8} {'人员':<12} {'任务数':>6} {'总工期(天)':>10} {'L3':>4} {'L4':>4} {'覆盖阶段':>10}")
    print(f"{'─' * 70}")

    sorted_roles = sorted(workload.items(), key=lambda x: x[1]["total_duration"], reverse=True)
    for role, stats in sorted_roles:
        person = team.get(role, "—")
        stages = ",".join(sorted(stats["stages"]))
        print(f"{role:<8} {person:<12} {stats['task_count']:>6} {stats['total_duration']:>10.0f} "
              f"{len(stats['l3_tasks']):>4} {len(stats['l4_tasks']):>4} {stages:>10}")

    # 负载均衡分析
    print(f"\n{'=' * 70}")
    print("⚖️ 负载均衡分析:")
    print(f"{'─' * 70}")

    durations = [s["total_duration"] for s in workload.values()]
    avg_duration = sum(durations) / len(durations) if durations else 0
    max_duration = max(durations) if durations else 0
    min_duration = min(durations) if durations else 0

    print(f"  平均工期负载: {avg_duration:.0f} 天")
    print(f"  最高工期负载: {max_duration:.0f} 天")
    print(f"  最低工期负载: {min_duration:.0f} 天")
    print(f"  负载差异比: {max_duration/min_duration:.1f}x" if min_duration > 0 else "")

    # 瓶颈识别
    print(f"\n{'=' * 70}")
    print("🔴 瓶颈与风险识别:")
    print(f"{'─' * 70}")

    for role, stats in sorted_roles:
        issues = []
        if stats["total_duration"] > avg_duration * 1.5:
            issues.append(f"工期负载超均值50%+（{stats['total_duration']:.0f} vs 均值{avg_duration:.0f}）")
        if len(stats["stages"]) >= 5:
            issues.append(f"跨{len(stats['stages'])}个阶段，上下文切换成本高")
        if stats["task_count"] > 30:
            issues.append(f"任务数过多（{stats['task_count']}个），管理幅度偏大")

        if issues:
            person = team.get(role, "—")
            print(f"\n  🔴 {role} ({person}):")
            for issue in issues:
                print(f"     → {issue}")

    # 阶段-角色热力图
    print(f"\n{'=' * 70}")
    print("🗺️ 阶段-角色参与矩阵:")
    print(f"{'─' * 70}")

    all_stages = sorted(set(s for stats in workload.values() for s in stats["stages"]))
    all_roles = [r for r, _ in sorted_roles]

    header = f"{'阶段':<6}" + "".join(f"{r:>8}" for r in all_roles)
    print(header)
    print(f"{'─' * 70}")

    for stage in all_stages:
        row = f"{stage:<6}"
        for role in all_roles:
            stage_tasks = [
                t for t in tasks
                if role in t["roles"] and any(
                    t["wbs"].startswith(stage) and (
                        len(t["wbs"]) == len(stage) or t["wbs"][len(stage)].isdigit()
                    )
                    for _ in [1]
                )
            ]
            count = len(stage_tasks)
            if count == 0:
                row += f"{'·':>8}"
            elif count <= 3:
                row += f"{'▪'+str(count):>8}"
            elif count <= 8:
                row += f"{'▪▪'+str(count):>8}"
            else:
                row += f"{'▪▪▪'+str(count):>8}"
        print(row)

    print(f"\n{'=' * 70}")
    print(f"分析完成。")
    return 0


if __name__ == "__main__":
    sys.exit(main())
